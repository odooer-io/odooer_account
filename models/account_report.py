# License: LGPL-3
import re
import datetime
import json
import logging
from collections import defaultdict
from dateutil.relativedelta import relativedelta

from odoo import api, fields, models, _
from odoo.exceptions import UserError
from odoo.tools import SQL
from odoo.tools.misc import format_date, formatLang

_logger = logging.getLogger(__name__)

# Regex to split account_codes formula on sign boundaries, e.g. "+100-200" → ["+100", "-200"]
_CODES_SPLIT_RE = re.compile(r'(?=[+-])')
# Each term: optional sign, account prefix, optional D/C balance character
_CODES_TERM_RE = re.compile(
    r'^(?P<sign>[+-]?)(?P<prefix>[A-Za-z0-9.]*)(?P<balance>[DC]?)$'
)


class AccountReport(models.Model):
    _inherit = 'account.report'

    # -------------------------------------------------------------------------
    # UI actions
    # -------------------------------------------------------------------------

    def action_open_report(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.client',
            'tag': 'odooer_account_report',
            'name': self.name,
            'context': {'report_id': self.id},
        }

    # -------------------------------------------------------------------------
    # Options helpers
    # -------------------------------------------------------------------------

    def _resolve_date_filter(self, filter_key, mode):
        """Compute (date_from_str, date_to_str) for a named date preset.

        Supported filter keys:
          this_year, last_year, this_month, last_month, this_quarter,
          today, YYYY (e.g. '2025'), YYYY-MM (e.g. '2026-01')
        Returns (date_from, date_to) as ISO strings, or (False, today) fallback.
        """
        today = fields.Date.context_today(self)
        company = self.env.company

        if filter_key == 'today':
            return (fields.Date.to_string(today) if mode == 'range' else False, fields.Date.to_string(today))

        if filter_key == 'this_year':
            fy = company.compute_fiscalyear_dates(today)
            d_from = fy['date_from']
            d_to = fy['date_to']
            if mode == 'single':
                return (False, fields.Date.to_string(d_to))
            return (fields.Date.to_string(d_from), fields.Date.to_string(d_to))

        if filter_key == 'last_year':
            last_year_date = today - relativedelta(years=1)
            fy = company.compute_fiscalyear_dates(last_year_date)
            d_from = fy['date_from']
            d_to = fy['date_to']
            if mode == 'single':
                return (False, fields.Date.to_string(d_to))
            return (fields.Date.to_string(d_from), fields.Date.to_string(d_to))

        if filter_key == 'this_month':
            d_from = today.replace(day=1)
            d_to = (d_from + relativedelta(months=1)) - datetime.timedelta(days=1)
            if mode == 'single':
                return (False, fields.Date.to_string(d_to))
            return (fields.Date.to_string(d_from), fields.Date.to_string(d_to))

        if filter_key == 'last_month':
            first_this = today.replace(day=1)
            last_month_end = first_this - datetime.timedelta(days=1)
            d_from = last_month_end.replace(day=1)
            d_to = last_month_end
            if mode == 'single':
                return (False, fields.Date.to_string(d_to))
            return (fields.Date.to_string(d_from), fields.Date.to_string(d_to))

        if filter_key == 'this_quarter':
            m = today.month
            q_start_month = ((m - 1) // 3) * 3 + 1
            d_from = today.replace(month=q_start_month, day=1)
            d_to = (d_from + relativedelta(months=3)) - datetime.timedelta(days=1)
            if mode == 'single':
                return (False, fields.Date.to_string(d_to))
            return (fields.Date.to_string(d_from), fields.Date.to_string(d_to))

        # YYYY-MM → specific month
        if filter_key and len(filter_key) == 7 and filter_key[4] == '-':
            try:
                year, month = int(filter_key[:4]), int(filter_key[5:])
                d_from = datetime.date(year, month, 1)
                d_to = (d_from + relativedelta(months=1)) - datetime.timedelta(days=1)
                if mode == 'single':
                    return (False, fields.Date.to_string(d_to))
                return (fields.Date.to_string(d_from), fields.Date.to_string(d_to))
            except (ValueError, TypeError):
                pass

        # YYYY → specific year (fiscal)
        if filter_key and len(filter_key) == 4 and filter_key.isdigit():
            try:
                year = int(filter_key)
                ref_date = datetime.date(year, 6, 15)  # mid-year to hit fiscal year
                fy = company.compute_fiscalyear_dates(ref_date)
                # If the fiscal year end is in a different year, try to match year-end year
                if fy['date_to'].year != year:
                    ref_date = datetime.date(year, 12, 31)
                    fy = company.compute_fiscalyear_dates(ref_date)
                d_from = fy['date_from']
                d_to = fy['date_to']
                if mode == 'single':
                    return (False, fields.Date.to_string(d_to))
                return (fields.Date.to_string(d_from), fields.Date.to_string(d_to))
            except (ValueError, TypeError):
                pass

        # custom / unknown: return what caller already has (no override)
        return None

    def _get_options(self, previous_options=None):
        """Return a full options dict for this report, merging user-supplied values
        from *previous_options* over the report's defaults."""
        self.ensure_one()
        options = {}

        # ── Date ──────────────────────────────────────────────────────────────
        today = fields.Date.context_today(self)
        mode = 'range' if self.filter_date_range else 'single'

        if previous_options and previous_options.get('date'):
            prev_date = dict(previous_options['date'])
            # If a named filter is set, re-resolve dates (so Today/This Month auto-advance)
            filter_key = prev_date.get('filter', 'custom')
            if filter_key and filter_key != 'custom':
                resolved = self._resolve_date_filter(filter_key, mode)
                if resolved:
                    prev_date['date_from'], prev_date['date_to'] = resolved
            options['date'] = prev_date
        else:
            default_filter = self.default_opening_date_filter or 'this_year'
            resolved = self._resolve_date_filter(default_filter, mode)
            if resolved:
                date_from, date_to = resolved
            else:
                year_start = today.replace(month=1, day=1)
                date_from = fields.Date.to_string(year_start) if mode == 'range' else False
                date_to = fields.Date.to_string(today)
            options['date'] = {
                'date_from': date_from,
                'date_to': date_to,
                'mode': mode,
                'filter': default_filter,
            }

        # ── Comparison ────────────────────────────────────────────────────────
        if previous_options and previous_options.get('comparison'):
            options['comparison'] = dict(previous_options['comparison'])
        else:
            options['comparison'] = {
                'enabled': False,
                'number_period': 1,
                'filter': 'previous_period',
                'periods': [],
            }

        if options['comparison']['enabled'] and self.filter_period_comparison:
            options['comparison']['periods'] = self._get_comparison_periods(options)

        # ── Display options ───────────────────────────────────────────────────
        options['show_draft'] = previous_options.get('show_draft', True) if previous_options else True
        options['hide_0_lines'] = previous_options.get('hide_0_lines', False) if previous_options else False
        options['unfold_all'] = previous_options.get('unfold_all', False) if previous_options else False
        options['unfolded_lines'] = previous_options.get('unfolded_lines', []) if previous_options else []

        # ── Report metadata ───────────────────────────────────────────────────
        options['report_id'] = self.id
        options['company_ids'] = [self.env.company.id]
        options['currency_id'] = self.env.company.currency_id.id

        # ── Columns (sent to JS so the header can be rendered) ────────────────
        options['columns'] = [
            {'id': col.id, 'name': col.name, 'expression_label': col.expression_label,
             'figure_type': col.figure_type or 'monetary'}
            for col in self.column_ids
        ]

        return options

    def _get_comparison_periods(self, options):
        """Generate comparison period dicts from the options comparison config."""
        periods = []
        date_from = fields.Date.from_string(options['date']['date_from'])
        date_to = fields.Date.from_string(options['date']['date_to'])
        n = options['comparison'].get('number_period', 1)
        filter_type = options['comparison'].get('filter', 'previous_period')

        for i in range(1, n + 1):
            if filter_type == 'previous_period':
                delta = date_to - date_from + datetime.timedelta(days=1)
                p_to = date_from - datetime.timedelta(days=1)
                p_from = p_to - delta + datetime.timedelta(days=1)
            elif filter_type == 'same_last_year':
                from dateutil.relativedelta import relativedelta
                p_from = date_from - relativedelta(years=i)
                p_to = date_to - relativedelta(years=i)
            else:
                break

            periods.append({
                'date_from': fields.Date.to_string(p_from),
                'date_to': fields.Date.to_string(p_to),
                'string': format_date(self.env, p_from) + ' – ' + format_date(self.env, p_to),
            })
            date_from = p_from
            date_to = p_to

        return periods

    # -------------------------------------------------------------------------
    # Public API — called by the OWL controller and HTTP endpoints
    # -------------------------------------------------------------------------

    def get_report_lines(self, options, offset=0):
        """Return paginated top-level report lines for the viewer.

        Returns::

            {
                'lines': [ <line_dict>, ... ],
                'load_more_offset': int,
                'load_more_remaining': int,
            }
        """
        self.ensure_one()
        all_values = self._get_all_line_values(options)

        top_lines = self.line_ids.filtered(lambda l: not l.parent_id).sorted('sequence')
        limit = self.load_more_limit or 0
        total = len(top_lines)

        if limit:
            page = top_lines[offset:offset + limit]
        else:
            page = top_lines

        lines = []
        for line in page:
            line_dict = self._build_line_dict(line, all_values, options)
            lines.append(line_dict)
            if options.get('unfold_all') or line.id in (options.get('unfolded_lines') or []):
                lines += self._get_expanded_lines(line, all_values, options)

        return {
            'lines': lines,
            'load_more_offset': offset + len(page),
            'load_more_remaining': max(0, total - offset - len(page)) if limit else 0,
        }

    def get_report_line_children(self, line_id, options):
        """Return child line dicts for drill-down expansion of *line_id*.

        For lines with ``groupby`` set, returns grouped AML summaries.
        For lines without ``groupby``, returns their sub-lines.
        """
        self.ensure_one()
        line = self.env['account.report.line'].browse(line_id)
        all_values = self._get_all_line_values(options)

        if line.groupby or line.user_groupby:
            return self._get_groupby_lines(line, options)
        return self._get_children_lines(line, all_values, options)

    # -------------------------------------------------------------------------
    # Audit / drill-down action
    # -------------------------------------------------------------------------

    def get_audit_action(self, line_id, options, audit_parent_line_id=None, audit_extra_domain=None):
        """Return an act_window action opening the AML journal items for *line_id*.

        For regular report lines the domain is built from the line's ``domain``
        engine expressions plus the current date/company filter.

        For groupby virtual sub-lines (e.g. a single Bank Account row expanded
        from "Bank and Cash") the caller also supplies *audit_parent_line_id*
        (the real report line) and *audit_extra_domain* (e.g.
        ``[('account_id', '=', 103)]``) to narrow the result.
        """
        self.ensure_one()
        parent_id = audit_parent_line_id or (int(line_id) if str(line_id).isdigit() else None)
        if not parent_id:
            return {}

        line = self.env['account.report.line'].browse(parent_id)
        domain = self._build_audit_domain_for_line(line, options)

        if audit_extra_domain:
            domain = audit_extra_domain + domain

        date_to = options['date']['date_to']
        date_from = options['date'].get('date_from')
        if date_from:
            date_label = f"{date_from} → {date_to}"
        else:
            date_label = f"As of {date_to}"

        return {
            'type': 'ir.actions.act_window',
            'name': f"{line.name} ({date_label})",
            'res_model': 'account.move.line',
            'view_mode': 'list,form',
            'views': [[False, 'list'], [False, 'form']],
            'domain': domain,
            'context': {'search_default_group_by_move': 1},
        }

    def _collect_domain_formulas(self, line):
        """Recursively collect all ``domain``-engine expression formulas from *line* and its
        descendants.  Parent/aggregation lines have no domain expressions themselves — we
        walk the entire sub-tree so clicking a parent total still drills down correctly.
        """
        formulas = [
            expr.formula
            for expr in line.expression_ids
            if expr.engine == 'domain'
        ]
        for child in line.children_ids:
            formulas += self._collect_domain_formulas(child)
        return formulas

    def _build_audit_domain_for_line(self, line, options):
        """Return the AML search domain for auditing *line* with *options*.

        Works for both leaf lines (own domain expressions) and parent/total lines
        (OR of all descendant leaf domains).
        """
        import ast

        date_from = options['date'].get('date_from')
        date_to = options['date']['date_to']
        show_draft = options.get('show_draft', False)
        company_ids = options.get('company_ids', [self.env.company.id])

        base = [
            ('company_id', 'in', company_ids),
            ('date', '<=', date_to),
            ('move_id.state', 'in', ['draft', 'posted'] if show_draft else ['posted']),
        ]
        if date_from:
            base.append(('date', '>=', date_from))

        formulas = self._collect_domain_formulas(line)
        parsed = []
        for formula in formulas:
            try:
                d = ast.literal_eval(formula)
                if d:
                    parsed.append(d)
            except Exception:
                _logger.warning('Could not parse audit domain: %s', formula)

        if not parsed:
            return base

        if len(parsed) == 1:
            return base + parsed[0]

        # OR together multiple sub-domains using Odoo's prefix notation:
        # 2 items: ['|', *d1, *d2]
        # 3 items: ['|', *d1, '|', *d2, *d3]  (chain: N-1 '|' operators interleaved)
        combined = []
        for i, d in enumerate(parsed):
            if i < len(parsed) - 1:
                combined.append('|')
            combined.extend(d)
        return base + combined

    # -------------------------------------------------------------------------

    def _get_expanded_lines(self, line, all_values, options):
        """Return sub-lines for *line* when it is in ``unfolded_lines``.

        Dispatches to ``_get_groupby_lines`` for groupby lines (virtual account/
        partner sub-rows) and to ``_get_children_lines`` for structural children.
        """
        if line.groupby or line.user_groupby:
            return self._get_groupby_lines(line, options)
        return self._get_children_lines(line, all_values, options)

    def _get_children_lines(self, parent_line, all_values, options):
        """Recursively build child line dicts for *parent_line*."""
        result = []
        for child in parent_line.children_ids.sorted('sequence'):
            result.append(self._build_line_dict(child, all_values, options))
            if options.get('unfold_all') or child.id in (options.get('unfolded_lines') or []):
                result += self._get_expanded_lines(child, all_values, options)
        return result
    def _build_line_dict(self, line, all_values, options):
        """Convert an account.report.line record into a serialisable dict."""
        columns = self._build_columns(line, all_values, options)

        # Determine whether all columns are zero (for hide_if_zero logic)
        all_zero = all(
            col.get('no_format') == 0 or col.get('no_format') is None
            for col in columns
        )
        if line.hide_if_zero and all_zero:
            return None  # callers must filter None

        has_children = bool(line.children_ids) or bool(line.groupby or line.user_groupby)

        return {
            'id': line.id,
            'name': line.name,
            'level': line.hierarchy_level,
            'code': line.code or '',
            'has_children': has_children,
            'unfoldable': has_children and line.foldable,
            'groupby': line.groupby or line.user_groupby or False,
            'columns': columns,
            'print_on_new_page': line.print_on_new_page,
        }

    def _build_columns(self, line, all_values, options):
        """Return the column value list for a line, including comparison periods."""
        columns = []
        line_vals = all_values.get(line.id, {})
        currency = self.env['res.currency'].browse(options['currency_id'])

        for col in self.column_ids.sorted('sequence'):
            value = line_vals.get(col.expression_label, 0.0)
            columns.append(self._format_column(value, col.figure_type, currency))

        # Append comparison period columns
        for period in options.get('comparison', {}).get('periods', []):
            period_values = all_values.get(f'{line.id}_cmp_{period["date_to"]}', {})
            for col in self.column_ids.sorted('sequence'):
                value = period_values.get(col.expression_label, 0.0)
                columns.append(self._format_column(value, col.figure_type, currency))

        return columns

    # -------------------------------------------------------------------------
    # Expression evaluation engine
    # -------------------------------------------------------------------------

    def _get_all_line_values(self, options):
        """Evaluate all lines and return ``{line_id: {label: value}}``."""
        self.ensure_one()
        all_lines = self.line_ids

        # First pass: evaluate account_codes and domain expressions (leaf nodes)
        results = {}
        for line in all_lines:
            results[line.id] = self._evaluate_line(line, options, results)

        # Second pass: sum_children and aggregation need a completed first pass
        for line in all_lines.sorted(lambda l: -l.hierarchy_level):
            for expr in line.expression_ids:
                if expr.engine in ('sum_children', 'aggregation'):
                    results.setdefault(line.id, {})
                    results[line.id][expr.label] = self._evaluate_expression(
                        expr, options, results
                    )

        # Comparison periods
        for period in options.get('comparison', {}).get('periods', []):
            period_opts = dict(options, date={
                'date_from': period['date_from'],
                'date_to': period['date_to'],
                'mode': 'range',
            }, comparison={'enabled': False, 'periods': []})
            for line in all_lines:
                key = f'{line.id}_cmp_{period["date_to"]}'
                results[key] = self._evaluate_line(line, period_opts, results)

        return results

    def _evaluate_line(self, line, options, existing_results):
        """Evaluate all expressions for *line* for the main period in *options*."""
        values = {}
        for expr in line.expression_ids:
            if expr.engine not in ('sum_children', 'aggregation'):
                values[expr.label] = self._evaluate_expression(expr, options, existing_results)
        return values

    def _evaluate_expression(self, expr, options, all_values=None):
        """Dispatch expression evaluation to the appropriate engine."""
        engine = expr.engine
        date_scope = expr.date_scope or 'strict_range'
        if engine == 'account_codes':
            return self._engine_account_codes(expr.formula, options, expr.subformula, date_scope=date_scope)
        elif engine == 'domain':
            return self._engine_domain(expr.formula, options, expr.subformula, date_scope=date_scope)
        elif engine == 'sum_children':
            return self._engine_sum_children(expr.report_line_id, options, all_values or {})
        elif engine == 'aggregation':
            return self._engine_aggregation(expr.formula, options, all_values or {})
        elif engine == 'external':
            return self._engine_external(expr, options)
        else:
            _logger.warning('Unsupported expression engine: %s', engine)
            return 0.0

    def _resolve_date_scope(self, options, date_scope):
        """Return (date_from, date_to) adjusted for the expression's date_scope.

        date_scope values:
          strict_range             - use options date exactly as-is
          from_beginning           - date_from=None (from beginning of time)
          from_fiscalyear          - date_from=fiscal year start, date_to unchanged
          to_beginning_of_fiscalyear - date_from=None, date_to=day before fiscal year start
          to_beginning_of_period   - date_from=None, date_to=options date_from - 1 day
        """
        date_from = options['date'].get('date_from') or None
        date_to = options['date']['date_to']

        if not date_scope or date_scope == 'strict_range':
            return date_from, date_to

        if date_scope == 'from_beginning':
            return None, date_to

        company = self.env.company
        ref_date = fields.Date.from_string(date_to) if isinstance(date_to, str) else date_to

        if date_scope == 'from_fiscalyear':
            fy = company.compute_fiscalyear_dates(ref_date)
            return fields.Date.to_string(fy['date_from']), date_to

        if date_scope == 'to_beginning_of_fiscalyear':
            fy = company.compute_fiscalyear_dates(ref_date)
            fy_start = fy['date_from']
            day_before = fy_start - datetime.timedelta(days=1)
            return None, fields.Date.to_string(day_before)

        if date_scope == 'to_beginning_of_period':
            if date_from:
                d = fields.Date.from_string(date_from) if isinstance(date_from, str) else date_from
                day_before = d - datetime.timedelta(days=1)
                return None, fields.Date.to_string(day_before)
            return None, date_to

        return date_from, date_to

    # ─── Engine: account_codes ─────────────────────────────────────────────────

    def _engine_account_codes(self, formula, options, subformula=None, date_scope=None):
        """Sum AML balances by account code prefixes.

        Formula syntax: ``[+-]<prefix>[D|C]`` terms joined without spaces.
        Example: ``+1-2`` → (sum of accounts 1xx) minus (sum of accounts 2xx)
        Example: ``4000C`` → credit side only of accounts 4000xx
        """
        eff_date_from, eff_date_to = self._resolve_date_scope(options, date_scope)
        total = 0.0
        for term_str in _CODES_SPLIT_RE.split(formula):
            if not term_str:
                continue
            m = _CODES_TERM_RE.match(term_str.strip())
            if not m:
                continue
            sign = -1.0 if m.group('sign') == '-' else 1.0
            prefix = m.group('prefix')
            balance_char = m.group('balance')  # 'D', 'C', or ''
            total += sign * self._sum_account_prefix(prefix, balance_char, options, eff_date_from, eff_date_to)
        return total

    def _sum_account_prefix(self, prefix, balance_char, options, date_from=None, date_to=None):
        """Execute a SQL query to sum AML amounts for accounts matching *prefix*.

        In Odoo 19, account.account.code is a computed field backed by
        code_store (jsonb keyed by company ID string, e.g. {"1": "101000"}).
        We query code_store->>'<company_id>' for the root company.
        """
        if date_from is None and date_to is None:
            date_from = options['date'].get('date_from')
            date_to = options['date']['date_to']
        company_ids = options.get('company_ids', [self.env.company.id])
        show_draft = options.get('show_draft', False)

        # Resolve root company IDs (code_store is keyed by root company id)
        root_company_id = str(self.env.company.root_id.id)

        states = ('draft', 'posted') if show_draft else ('posted',)
        state_placeholders = ','.join(['%s'] * len(states))

        if balance_char == 'D':
            amount_expr = 'SUM(aml.debit)'
        elif balance_char == 'C':
            amount_expr = 'SUM(aml.credit)'
        else:
            amount_expr = 'SUM(aml.balance)'

        params = list(states) + [tuple(company_ids)]

        if prefix:
            # code_store is jsonb: {"<company_id>": "code_value"}
            where_prefix = f"AND (acc.code_store->>%s) LIKE %s"
            params += [root_company_id, prefix + '%']
        else:
            where_prefix = ''

        if date_from:
            where_date = "AND am.date >= %s AND am.date <= %s"
            params += [date_from, date_to]
        else:
            where_date = "AND am.date <= %s"
            params.append(date_to)

        query = f"""
            SELECT COALESCE({amount_expr}, 0.0)
              FROM account_move_line aml
              JOIN account_move am ON aml.move_id = am.id
              JOIN account_account acc ON aml.account_id = acc.id
             WHERE am.state IN ({state_placeholders})
               AND aml.company_id IN %s
               {where_prefix}
               {where_date}
        """
        self.env.cr.execute(query, params)
        return self.env.cr.fetchone()[0] or 0.0

    # ─── Engine: domain ───────────────────────────────────────────────────────

    def _engine_domain(self, formula, options, subformula=None, date_scope=None):
        """Evaluate a domain expression against account.move.line.

        subformula values:
          'sum' or '' or None  → SUM(balance)
          '-sum'               → -SUM(balance)   (for credit-normal accounts like income)
          'debit'              → SUM(debit)
          'credit'             → SUM(credit)
          'sum_abs'            → ABS(SUM(balance))
        """
        try:
            import ast
            domain = ast.literal_eval(formula)
        except Exception:
            _logger.warning('Could not parse domain formula: %s', formula)
            return 0.0

        date_from, date_to = self._resolve_date_scope(options, date_scope)
        show_draft = options.get('show_draft', False)
        company_ids = options.get('company_ids', [self.env.company.id])

        base_domain = [
            ('company_id', 'in', company_ids),
            ('date', '<=', date_to),
            ('move_id.state', 'in', ['draft', 'posted'] if show_draft else ['posted']),
        ]
        if date_from:
            base_domain.append(('date', '>=', date_from))

        full_domain = base_domain + domain
        amls = self.env['account.move.line'].search(full_domain)

        sf = (subformula or '').strip()
        if sf == '-sum':
            return -sum(amls.mapped('balance'))
        elif sf == 'debit':
            return sum(amls.mapped('debit'))
        elif sf == 'credit':
            return sum(amls.mapped('credit'))
        elif sf == 'sum_abs':
            return abs(sum(amls.mapped('balance')))
        return sum(amls.mapped('balance'))

    # ─── Engine: sum_children ─────────────────────────────────────────────────

    def _engine_sum_children(self, line, options, all_values):
        """Sum the 'balance' expression values of all direct children."""
        total = 0.0
        for child in line.children_ids:
            child_val = all_values.get(child.id, {})
            total += child_val.get('balance', 0.0)
        return total

    # ─── Engine: aggregation ──────────────────────────────────────────────────

    def _engine_aggregation(self, formula, options, all_values):
        """Evaluate an aggregation formula referencing other line codes.

        Supports: ``CODE1.balance + CODE2.balance - CODE3.balance``
        and the hard formula ``sum_children`` (handled separately).
        """
        # Build a lookup: code → line_id
        code_map = {
            line.code: line.id
            for line in self.line_ids
            if line.code
        }

        def resolve_ref(match):
            code, label = match.group(1), match.group(2)
            line_id = code_map.get(code)
            if line_id is None:
                return '0.0'
            return str(all_values.get(line_id, {}).get(label, 0.0))

        expr_str = re.sub(
            r'([A-Za-z0-9_]+)\.([A-Za-z0-9_]+)',
            resolve_ref,
            formula,
        )
        try:
            return float(eval(expr_str, {"__builtins__": {}}))  # noqa: S307
        except Exception:
            _logger.warning('Could not evaluate aggregation formula: %s → %s', formula, expr_str)
            return 0.0

    # ─── Engine: external ─────────────────────────────────────────────────────

    def _engine_external(self, expr, options):
        """Return the most-recent external value for this expression, if any."""
        domain = [
            ('report_line_id', '=', expr.report_line_id.id),
            ('target_date', '<=', options['date']['date_to']),
            ('company_id', 'in', options.get('company_ids', [self.env.company.id])),
        ]
        rec = self.env['account.report.external.value'].search(
            domain, order='target_date desc', limit=1
        )
        return rec.value if rec else 0.0

    # ─── Groupby lines ────────────────────────────────────────────────────────

    def _get_groupby_lines(self, line, options):
        """Generate grouped sub-lines for a line that has groupby set."""
        groupby_fields = [
            f.strip()
            for f in (line.user_groupby or line.groupby or '').split(',')
            if f.strip()
        ]
        if not groupby_fields:
            return []

        date_from = options['date'].get('date_from')
        date_to = options['date']['date_to']
        show_draft = options.get('show_draft', False)
        company_ids = options.get('company_ids', [self.env.company.id])

        domain = [
            ('company_id', 'in', company_ids),
            ('date', '<=', date_to),
            ('move_id.state', 'in', ['draft', 'posted'] if show_draft else ['posted']),
        ]
        if date_from:
            domain.append(('date', '>=', date_from))

        # Apply expression domain filters from the line's expressions
        for expr in line.expression_ids:
            if expr.engine == 'domain':
                try:
                    import ast
                    domain += ast.literal_eval(expr.formula)
                except Exception:
                    pass

        amls = self.env['account.move.line'].search(domain)

        # Group in Python; for large datasets a SQL GROUP BY would be preferable
        groups = defaultdict(float)
        group_names = {}
        for aml in amls:
            key_parts = []
            name_parts = []
            for gf in groupby_fields:
                val = aml[gf] if gf in aml._fields else None
                if hasattr(val, 'id'):
                    key_parts.append(val.id)
                    name_parts.append(val.display_name or str(val.id))
                else:
                    key_parts.append(val)
                    name_parts.append(str(val) if val is not None else _('Undefined'))
            key = tuple(key_parts)
            groups[key] += aml.balance
            group_names[key] = ' / '.join(name_parts)

        currency = self.env['res.currency'].browse(options['currency_id'])
        result = []
        for idx, (key, balance) in enumerate(sorted(groups.items(), key=lambda x: -abs(x[1]))):
            # Build extra domain to drill down to just this groupby bucket
            extra_domain = []
            for i, gf in enumerate(groupby_fields):
                v = key[i]
                if v is not None:
                    extra_domain.append((gf, '=', v))

            result.append({
                'id': f'groupby_{line.id}_{idx}',
                'name': group_names[key],
                'level': line.hierarchy_level + 2,
                'code': '',
                'has_children': False,
                'unfoldable': False,
                'groupby': False,
                'columns': [self._format_column(balance, 'monetary', currency)],
                'print_on_new_page': False,
                'audit_parent_line_id': line.id,
                'audit_extra_domain': extra_domain,
            })
        return result

    # -------------------------------------------------------------------------
    # Value formatting
    # -------------------------------------------------------------------------

    def _format_column(self, value, figure_type, currency=None):
        """Return a column cell dict with formatted string and raw value."""
        if figure_type == 'monetary':
            no_format = round(value, currency.decimal_places if currency else 2) if value else 0.0
            name = formatLang(self.env, no_format, currency_obj=currency) if currency else str(no_format)
        elif figure_type == 'percentage':
            no_format = round(value * 100, 2) if value else 0.0
            name = f'{no_format:.2f}%'
        elif figure_type == 'integer':
            no_format = int(value) if value else 0
            name = str(no_format)
        elif figure_type == 'float':
            no_format = round(value, 4) if value else 0.0
            name = str(no_format)
        else:
            no_format = value
            name = str(value) if value is not None else ''

        return {
            'name': name,
            'no_format': no_format,
            'figure_type': figure_type,
            'class': 'number' if figure_type in ('monetary', 'percentage', 'integer', 'float') else '',
        }

    # -------------------------------------------------------------------------
    # Export helpers
    # -------------------------------------------------------------------------

    def get_xlsx(self, options):
        """Return an XLSX bytes object for the current report."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError(_('Please install the openpyxl library to export to XLSX.'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.name[:31]

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='2F5496')
        number_alignment = Alignment(horizontal='right')

        # Header row
        headers = [_('Name')] + [col.name for col in self.column_ids.sorted('sequence')]
        for period in options.get('comparison', {}).get('periods', []):
            headers += [f'{col.name} ({period["string"]})' for col in self.column_ids.sorted('sequence')]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        data = self.get_report_lines(options)
        row_idx = 2
        for line in data['lines']:
            if line is None:
                continue
            indent = '    ' * max(0, (line['level'] - 1) // 2)
            ws.cell(row=row_idx, column=1, value=indent + line['name'])
            if line['level'] == 1:
                ws.cell(row=row_idx, column=1).font = Font(bold=True)

            for col_idx, col_data in enumerate(line['columns'], 2):
                cell = ws.cell(row=row_idx, column=col_idx, value=col_data.get('no_format'))
                if col_data.get('class') == 'number':
                    cell.alignment = number_alignment
            row_idx += 1

        # Auto-size columns
        for col_idx in range(1, len(headers) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].auto_size = True

        import io
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
